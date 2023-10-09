import json
import os
import datetime
import sys

import openpyxl
from django.core.exceptions import FieldDoesNotExist
from django.db.models import Q, Count
from django.http import HttpResponse
from django.utils import timezone

from R4C.settings import BASE_DIR
from robots.models import Robot

sys.path.append('/Library/Frameworks/Python.framework/Versions/3.10/lib/python3.10/site-packages')

# Create your views here.
def manufactured_robot(request):
    if request.method == 'POST':
        input_data = json.loads(request.body)
    for key, value in input_data.items():
        try:
            Robot._meta.get_field(key)
        except FieldDoesNotExist:
            return HttpResponse('Входные данные не соотвествуют модели.')
    serial = f'{input_data["model"]} - {input_data["version"]}'
    new_robot = Robot.objects.create(serial=serial, model=input_data["model"], version=input_data["version"], created=input_data["created"])
    new_robot.save()
    return HttpResponse(new_robot)

def robots_report(request):
    robots = Robot.objects.filter(created__gte=(timezone.now() - datetime.timedelta(days=10)))
    report = openpyxl.Workbook()
    filename = "weekly_report.xlsx"
    robots_serials = {}
    for robot in robots:
        robot_model = robots.filter(Q(model__exact=robot.model))
        robot_version = robot_model.filter(Q(version__exact=robot.version)).aggregate(Count("version"))
        robots_serials[robot.model] = {
            model.version: robot_version.get("version__count") for model in
            robot_model}
    for outer_k, outer_v in robots_serials.items():
        sheet = report.create_sheet(outer_k)
        sheet['A1'] = 'Модель'
        sheet['B1'] = 'Версия'
        sheet.merge_cells('C1:E1')
        sheet['C1'] = 'Количество за неделю'
        n = 2
        for inner_k, inner_v in outer_v.items():
            sheet[f'A{n}'] = outer_k
            sheet[f'B{n}'] = inner_k
            sheet.merge_cells(f'C{n}:E{n}')
            sheet[f'C{n}'] = inner_v
            n += 1
        report.save(filename)
    filepath = os.path.join(BASE_DIR, filename)
    with open(filepath, 'rb') as rep:
        response = HttpResponse(rep.read(), content_type="application/vnd.ms-excel")
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response
